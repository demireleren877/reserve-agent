"""Ham hasar verisi parser — inspect + mapping destekli."""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any


@dataclass
class ClaimRecord:
    dosya_no: str
    brans: str
    hasar_tarihi: date
    gelisim_tarihi: date
    odeme: float
    muallak: float


# Zorunlu alan adları
REQUIRED_FIELDS = ["dosya_no", "brans", "hasar_tarihi", "gelisim_tarihi", "odeme", "muallak"]

# Otomatik tahmin için alias → field
_AUTO_ALIASES: dict[str, str] = {
    "dosya_no": "dosya_no", "dosyano": "dosya_no", "file_no": "dosya_no",
    "fileno": "dosya_no", "claim_no": "dosya_no", "claimno": "dosya_no",
    "hasar_no": "dosya_no", "hasarno": "dosya_no",
    "brans": "brans", "branch": "brans", "lob": "brans", "line_of_business": "brans",
    "hasar_tarihi": "hasar_tarihi", "hasartarihi": "hasar_tarihi",
    "loss_date": "hasar_tarihi", "lossdate": "hasar_tarihi",
    "accident_date": "hasar_tarihi", "accidentdate": "hasar_tarihi",
    "gelisim_tarihi": "gelisim_tarihi", "gelisimtarihi": "gelisim_tarihi",
    "dev_date": "gelisim_tarihi", "devdate": "gelisim_tarihi",
    "development_date": "gelisim_tarihi", "report_date": "gelisim_tarihi",
    "reportdate": "gelisim_tarihi", "valuation_date": "gelisim_tarihi",
    "valuationdate": "gelisim_tarihi",
    "odeme": "odeme", "payment": "odeme", "paid": "odeme", "odenen": "odeme",
    "muallak": "muallak", "outstanding": "muallak", "reserve": "muallak",
    "case_reserve": "muallak", "casereserve": "muallak", "ibnr": "muallak",
}

_DATE_FORMATS = [
    "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y.%m.%d",
    "%d.%m.%y", "%d/%m/%y", "%m/%d/%Y", "%Y%m%d",
    # datetime with time component
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M",
    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
]

# Çeyrek → ayın ilk günü (yyyyqq: 200001→2000-Q1, 200004→2000-Q4)
_QUARTER_START_MONTH = {1: 1, 2: 4, 3: 7, 4: 10}


def _try_parse_yyyyqq(s: str) -> date | None:
    """
    yyyyqq formatı: 6 haneli tam sayı, son 2 hane çeyrek (01-04).
    Örn: 200001 → 2000-01-01, 200304 → 2003-10-01.
    Ayrıca 'yyyyQq' (2000Q1, 2000Q4) formatını da destekler.
    """
    s = s.strip()
    # "2000Q1" / "2000q4" formatı
    import re as _re
    m = _re.fullmatch(r"(\d{4})[Qq]([1-4])", s)
    if m:
        year, q = int(m.group(1)), int(m.group(2))
        return date(year, _QUARTER_START_MONTH[q], 1)
    # "200001" … "200004" — 6 haneli
    if _re.fullmatch(r"\d{6}", s):
        year, q = int(s[:4]), int(s[4:])
        if 1 <= q <= 4:
            return date(year, _QUARTER_START_MONTH[q], 1)
    return None


def _normalize_col(name: str) -> str:
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))
    name = name.lower().strip()
    return re.sub(r"[^a-z0-9]+", "_", name).strip("_")


def _auto_suggest(headers: list[str]) -> dict[str, str]:
    """Header listesinden field → column_name tahmini üret."""
    suggestion: dict[str, str] = {}
    for h in headers:
        key = _normalize_col(h)
        field = _AUTO_ALIASES.get(key)
        if field and field not in suggestion:
            suggestion[field] = h
    return suggestion


def _parse_date(s: str) -> date:
    s = s.strip()
    if not s:
        raise ValueError("Tarih boş olamaz")
    # yyyyqq / yyyyQq kontrolü (sayısal veya Q harfli)
    qqdate = _try_parse_yyyyqq(s)
    if qqdate is not None:
        return qqdate
    # Excel serial number (5-6 haneli tam sayı, çeyrek aralığının dışında)
    if s.isdigit():
        n = int(s)
        if 20000 < n < 100000:
            return date(1899, 12, 30) + timedelta(days=n)
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Tarih formatı tanınamadı: {s!r}")


def _parse_float(s: str) -> float:
    if not s or s.strip() in ("-", ""):
        return 0.0
    s = s.strip().replace("\xa0", "").replace(" ", "")
    if "," in s and "." in s:
        if s.rindex(",") > s.rindex("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    return float(s)


def _detect_delimiter(sample: str) -> str:
    for delim in (";", ",", "\t", "|"):
        if delim in sample:
            return delim
    return ","


def _is_excel(content: bytes, filename: str) -> bool:
    return (
        content[:4] in (b"PK\x03\x04", b"\xd0\xcf\x11\xe0")
        or filename.lower().endswith((".xlsx", ".xls"))
    )


# ─── Inspect ─────────────────────────────────────────────────────────────────

def inspect_file(content: bytes, filename: str) -> dict[str, Any]:
    """
    Dosyayı açıp sheet listesi ve her sheet'in başlıklarını döndürür.
    CSV için sheets=[None] olur.
    Dönen yapı:
      {
        "sheets": ["Sheet1", "Sheet2"] | [null],  # null = CSV (sheet yok)
        "headers": {"Sheet1": ["A","B",...], null: [...]} ,
        "preview": {"Sheet1": [[...],[...]], ...}   # ilk 5 satır (header hariç)
        "suggested_mapping": {"Sheet1": {"dosya_no": "Dosya No", ...}, ...}
      }
    """
    if _is_excel(content, filename):
        return _inspect_excel(content)
    return _inspect_csv(content)


def _inspect_excel(content: bytes) -> dict[str, Any]:
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = wb.sheetnames
    headers_map: dict[str, list[str]] = {}
    preview_map: dict[str, list[list[str]]] = {}
    suggested_map: dict[str, dict[str, str]] = {}

    for sheet_name in sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True, max_row=6))
        if not rows:
            headers_map[sheet_name] = []
            preview_map[sheet_name] = []
            suggested_map[sheet_name] = {}
            continue
        hdrs = [str(c) if c is not None else "" for c in rows[0]]
        headers_map[sheet_name] = hdrs
        preview_map[sheet_name] = [
            [str(c) if c is not None else "" for c in row]
            for row in rows[1:]
            if any(c is not None for c in row)
        ]
        suggested_map[sheet_name] = _auto_suggest(hdrs)

    wb.close()
    return {
        "sheets": sheets,
        "headers": headers_map,
        "preview": preview_map,
        "suggested_mapping": suggested_map,
    }


def _inspect_csv(content: bytes) -> dict[str, Any]:
    text = content.decode("utf-8-sig", errors="replace")
    delim = _detect_delimiter(text[:2000])
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(reader)
    if not rows:
        return {"sheets": [None], "headers": {None: []}, "preview": {None: []}, "suggested_mapping": {None: {}}}

    hdrs = rows[0]
    preview = [r for r in rows[1:6] if any(c.strip() for c in r)]
    return {
        "sheets": [None],
        "headers": {None: hdrs},
        "preview": {None: preview},
        "suggested_mapping": {None: _auto_suggest(hdrs)},
    }


# ─── Parse with explicit mapping ─────────────────────────────────────────────

def parse_with_mapping(
    content: bytes,
    filename: str,
    column_mapping: dict[str, str],  # field → column_name
    sheet_name: str | None = None,
) -> list[ClaimRecord]:
    """
    Kullanıcının belirlediği column_mapping ile dosyayı parse eder.
    column_mapping örneği: {"dosya_no": "Poliçe No", "brans": "Branş", ...}
    """
    missing = [f for f in REQUIRED_FIELDS if f not in column_mapping or not column_mapping[f]]
    if missing:
        raise ValueError(f"Eşleştirilmemiş alanlar: {', '.join(missing)}")

    if _is_excel(content, filename):
        return _parse_excel_mapped(content, sheet_name, column_mapping)
    return _parse_csv_mapped(content, column_mapping)


def _parse_excel_mapped(
    content: bytes,
    sheet_name: str | None,
    column_mapping: dict[str, str],
) -> list[ClaimRecord]:
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Excel sayfası boş")

    headers = [str(c) if c is not None else "" for c in rows[0]]
    idx_map = _build_index_map(headers, column_mapping)
    records = []

    for row_num, row in enumerate(rows[1:], start=2):
        row_s = [str(c) if c is not None else "" for c in row]
        if not any(row_s):
            continue
        try:
            records.append(_extract(row_s, idx_map))
        except ValueError as e:
            raise ValueError(f"Satır {row_num}: {e}") from e

    return records


def _parse_csv_mapped(
    content: bytes,
    column_mapping: dict[str, str],
) -> list[ClaimRecord]:
    text = content.decode("utf-8-sig", errors="replace")
    delim = _detect_delimiter(text[:2000])
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(reader)

    if not rows:
        raise ValueError("CSV dosyası boş")

    idx_map = _build_index_map(rows[0], column_mapping)
    records = []

    for row_num, row in enumerate(rows[1:], start=2):
        if not any(c.strip() for c in row):
            continue
        try:
            records.append(_extract(row, idx_map))
        except ValueError as e:
            raise ValueError(f"Satır {row_num}: {e}") from e

    return records


def _build_index_map(headers: list[str], column_mapping: dict[str, str]) -> dict[str, int]:
    """column_mapping (field→col_name) → field→col_index."""
    header_index = {h: i for i, h in enumerate(headers)}
    result: dict[str, int] = {}
    for field, col_name in column_mapping.items():
        if col_name not in header_index:
            raise ValueError(f"'{col_name}' sütunu dosyada bulunamadı")
        result[field] = header_index[col_name]
    return result


def _extract(row: list[str], idx_map: dict[str, int]) -> ClaimRecord:
    def get(field: str) -> str:
        i = idx_map[field]
        return row[i] if i < len(row) else ""

    return ClaimRecord(
        dosya_no=get("dosya_no").strip(),
        brans=get("brans").strip(),
        hasar_tarihi=_parse_date(get("hasar_tarihi")),
        gelisim_tarihi=_parse_date(get("gelisim_tarihi")),
        odeme=_parse_float(get("odeme")),
        muallak=_parse_float(get("muallak")),
    )
