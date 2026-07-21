"""Excel parser — uzun (tidy) format ve klasik pivot üçgen formatını destekler.

Uzun format (tercih edilen):
    ACCIDENT_YEAR | DEVELOPMENT_DATE | PAID
    2020          | 2020             | 1000
    2020          | 2021             | 1500
    ...

Kolon adları esnek: accident/origin/kaza, development/dev/gelişim, paid/incurred/value/tutar.
Dönemler yıllık (2020) veya çeyreklik (2020Q1, 2020-Q1, 2020.1) olabilir.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

# openpyxl LAZY import edilir (startup'ı yavaşlatmasın; sadece Excel parse edilince
# yüklenir — kullanıcı giriş yaptıktan sonra veri yüklerken).

from app.core.triangle import Granularity, Triangle, TriangleType


class ParseError(ValueError):
    pass


ORIGIN_ALIASES = (
    "accident_year", "accident_date", "accident_period", "accident_quarter",
    "accident", "origin_year", "origin_period", "origin",
    "kaza_yili", "kaza", "donem",
)
DEV_ALIASES = (
    "development_date", "development_period", "development_year",
    "development_quarter", "development", "dev_period", "dev",
    "gelisim", "gelisim_donemi",
)
VALUE_ALIASES = (
    "paid", "incurred", "value", "amount", "loss", "tutar", "odenen", "tahakkuk",
)
DOSYA_ALIASES = (
    "dosya_no", "dosya", "file_no", "file", "claim_id", "policy_no", "police_no",
)
PREMIUM_ALIASES = (
    "premium", "earned_premium", "gross_premium", "written_premium",
    "kazanilan_prim", "prim", "kazanilan", "yazilan_prim",
)


@dataclass
class ParseOptions:
    triangle_type: TriangleType = TriangleType.PAID
    origin_granularity: Granularity = Granularity.YEARLY
    development_granularity: Granularity = Granularity.YEARLY
    cumulative: bool = True
    sheet_name: str | None = None


def parse_premiums_from_excel(
    content: bytes,
    origin_granularity: Granularity = Granularity.YEARLY,
    sheet_name: str | None = None,
) -> dict[str, float]:
    """Parse (origin, premium) long-format Excel → {origin_label: premium}."""
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
    except (InvalidFileException, Exception) as e:
        raise ParseError(f"Excel okunamadı: {e}") from e

    ws = wb[sheet_name] if sheet_name else wb.active
    if ws is None:
        raise ParseError("Sheet bulunamadı")

    rows: list[list[Any]] = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        raise ParseError("Sheet boş")

    header_idx: int | None = None
    for idx, row in enumerate(rows):
        names = [_normalize(c) for c in row]
        if (
            _find_match(names, ORIGIN_ALIASES) is not None
            and _find_match(names, PREMIUM_ALIASES) is not None
        ):
            header_idx = idx
            break
    if header_idx is None:
        raise ParseError(
            "Başlık bulunamadı. Kolonlar: ACCIDENT_YEAR (veya origin/kaza) + "
            "PREMIUM (veya kazanılan_prim)."
        )

    header = rows[header_idx]
    names = [_normalize(c) for c in header]
    origin_col = _find_match(names, ORIGIN_ALIASES)
    premium_col = _find_match(names, PREMIUM_ALIASES)
    assert origin_col is not None and premium_col is not None

    out: dict[str, float] = {}
    for row in rows[header_idx + 1:]:
        if _is_blank_row(row):
            continue
        if origin_col >= len(row) or premium_col >= len(row):
            continue
        raw_origin = row[origin_col]
        raw_prem = row[premium_col]
        if raw_origin in (None, "") or raw_prem in (None, ""):
            continue
        try:
            origin_label, _ = _parse_period(raw_origin, origin_granularity)
        except ValueError as e:
            raise ParseError(f"Origin çözümlenemedi: {raw_origin!r} ({e})") from e
        val = _to_float(raw_prem)
        if val is None:
            continue
        out[origin_label] = out.get(origin_label, 0.0) + val

    if not out:
        raise ParseError("Hiç prim kaydı bulunamadı")
    return out


def parse_triangle_from_excel(
    content: bytes,
    options: ParseOptions | None = None,
) -> tuple[Triangle, dict | None]:
    opts = options or ParseOptions()
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
    except (InvalidFileException, Exception) as e:
        raise ParseError(f"Excel dosyası okunamadı: {e}") from e

    ws = wb[opts.sheet_name] if opts.sheet_name else wb.active
    if ws is None:
        raise ParseError("Sheet bulunamadı")

    rows: list[list[Any]] = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        raise ParseError("Sheet boş")

    header_idx = _find_long_header(rows)
    if header_idx is not None:
        return _parse_long_format(rows, header_idx, opts)

    pivot_idx = _find_pivot_header(rows)
    if pivot_idx is not None:
        return _parse_pivot_format(rows, pivot_idx, opts), None

    raise ParseError(
        "Başlık bulunamadı. Uzun format için ACCIDENT_YEAR / DEVELOPMENT_DATE / PAID "
        "gibi kolonlar; pivot için sol sütun origin, üst satır development olmalı."
    )


def _find_long_header(rows: list[list[Any]]) -> int | None:
    for idx, row in enumerate(rows):
        names = [_normalize(c) for c in row]
        if _find_match(names, ORIGIN_ALIASES) is not None \
           and _find_match(names, DEV_ALIASES) is not None \
           and _find_match(names, VALUE_ALIASES) is not None:
            return idx
    return None


def _find_pivot_header(rows: list[list[Any]]) -> int | None:
    for idx, row in enumerate(rows):
        if not row or len(row) < 2:
            continue
        first = row[0]
        if first is None or first == "":
            continue
        if isinstance(first, int | float) and 1900 <= float(first) <= 2100:
            continue
        numeric_count = sum(
            1 for cell in row[1:]
            if isinstance(cell, int | float) and 0 < float(cell) <= 60
        )
        if numeric_count >= 2:
            return idx
    return None


def _parse_long_format(
    rows: list[list[Any]], header_idx: int, opts: ParseOptions
) -> tuple[Triangle, dict | None]:
    header = rows[header_idx]
    names = [_normalize(c) for c in header]
    origin_col = _find_match(names, ORIGIN_ALIASES)
    dev_col = _find_match(names, DEV_ALIASES)
    value_col = _find_match(names, VALUE_ALIASES)
    dosya_col = _find_match(names, DOSYA_ALIASES)
    assert origin_col is not None and dev_col is not None and value_col is not None

    records: list[tuple[str, int, float]] = []
    origin_order: dict[str, int] = {}
    max_dev_rank: int | None = None
    # dosya records: (origin_label, dev_label, dosya_no, value)
    dosya_records: list[tuple[str, str, str, float]] = []

    for row in rows[header_idx + 1:]:
        if _is_blank_row(row):
            continue
        if origin_col >= len(row) or dev_col >= len(row) or value_col >= len(row):
            continue
        raw_origin = row[origin_col]
        raw_dev = row[dev_col]
        raw_val = row[value_col]
        if raw_origin in (None, "") or raw_dev in (None, ""):
            continue

        try:
            origin_label, origin_rank = _parse_period(raw_origin, opts.origin_granularity)
        except ValueError as e:
            raise ParseError(f"Origin değeri çözümlenemedi: {raw_origin!r} ({e})") from e
        try:
            dev_label, dev_rank = _parse_period(raw_dev, opts.development_granularity)
        except ValueError as e:
            raise ParseError(f"Development değeri çözümlenemedi: {raw_dev!r} ({e})") from e

        age = dev_rank - origin_rank
        if age < 0:
            raise ParseError(
                f"Development ({raw_dev}) origin'den ({raw_origin}) önce olamaz"
            )

        val = _to_float(raw_val)
        if val is None:
            continue

        if origin_label not in origin_order:
            origin_order[origin_label] = origin_rank
        if max_dev_rank is None or dev_rank > max_dev_rank:
            max_dev_rank = dev_rank
        records.append((origin_label, age, val))

        if dosya_col is not None and dosya_col < len(row):
            raw_dosya = row[dosya_col]
            if raw_dosya not in (None, ""):
                dosya_no = str(raw_dosya).strip()
                dosya_records.append((origin_label, dev_label, dosya_no, val))

    if not records or max_dev_rank is None:
        raise ParseError("Uzun formatta veri satırı bulunamadı")

    # Age is currently in quarter units (consistent across granularities).
    # When both granularities are yearly, collapse to year units for cleaner display.
    both_yearly = (
        opts.origin_granularity == Granularity.YEARLY
        and opts.development_granularity == Granularity.YEARLY
    )
    if both_yearly:
        records = [(o, a // 4, v) for (o, a, v) in records]

    origins = sorted(origin_order.keys(), key=lambda k: origin_order[k])

    # Report rank dictates the rightmost observable dev date; each origin's
    # observable ages = 0..(report_rank - origin_rank). Normalize to display unit.
    if both_yearly:
        origin_max_age = {
            o: (max_dev_rank - origin_order[o]) // 4 for o in origins
        }
    else:
        origin_max_age = {o: max_dev_rank - origin_order[o] for o in origins}

    overall_max_age = max(origin_max_age.values())
    ages = list(range(0, overall_max_age + 1))

    agg: dict[tuple[str, int], float] = {}
    for o, a, v in records:
        key = (o, a)
        agg[key] = agg.get(key, 0.0) + v

    values: list[list[float | None]] = [[None] * len(ages) for _ in origins]

    for i, o in enumerate(origins):
        max_age = origin_max_age[o]
        for a in range(max_age + 1):
            v = agg.get((o, a))
            if v is not None:
                values[i][a] = v
            elif not opts.cumulative:
                # Zero payment in that period — fill so running sum is continuous.
                values[i][a] = 0.0
            # cumulative mode: leave None and carry-forward below

    if opts.cumulative:
        # Carry-forward gaps within observable range for cumulative inputs.
        for i, o in enumerate(origins):
            max_age = origin_max_age[o]
            last: float | None = None
            for a in range(max_age + 1):
                if values[i][a] is not None:
                    last = values[i][a]
                elif last is not None:
                    values[i][a] = last
    else:
        # Incremental → running sum within observable range.
        for i, o in enumerate(origins):
            max_age = origin_max_age[o]
            running = 0.0
            for a in range(max_age + 1):
                cell = values[i][a]
                if cell is None:
                    continue
                running += cell
                values[i][a] = running

    try:
        triangle = Triangle(
            origin_periods=origins,
            development_periods=ages,
            values=values,
            triangle_type=opts.triangle_type,
            origin_granularity=opts.origin_granularity,
            development_granularity=opts.development_granularity,
        )
    except ValueError as e:
        raise ParseError(f"Üçgen doğrulaması başarısız: {e}") from e

    file_data: dict | None = None
    if dosya_records:
        fd: dict[str, dict[str, dict[str, float]]] = {}
        for origin_label, dev_label, dosya_no, val in dosya_records:
            fd.setdefault(origin_label, {}).setdefault(dev_label, {})[dosya_no] = (
                fd.get(origin_label, {}).get(dev_label, {}).get(dosya_no, 0.0) + val
            )
        file_data = fd

    return triangle, file_data


def _parse_pivot_format(
    rows: list[list[Any]], header_idx: int, opts: ParseOptions
) -> Triangle:
    header = rows[header_idx]
    development_periods: list[int] = []
    for cell in header[1:]:
        if cell is None or cell == "":
            continue
        try:
            development_periods.append(int(cell))
        except (TypeError, ValueError) as e:
            raise ParseError(
                f"Development period başlığı sayısal olmalı: {cell!r}"
            ) from e
    if not development_periods:
        raise ParseError("Başlık satırında development period yok")
    # Normalize to 0-indexed so pivot [1,2,...,N] and long-format [0,1,...,N-1]
    # produce the same development_periods array — cdfInitial keys stay consistent.
    min_dev = min(development_periods)
    development_periods = [d - min_dev for d in development_periods]
    n_dev = len(development_periods)

    origin_periods: list[str] = []
    values: list[list[float | None]] = []

    for row in rows[header_idx + 1:]:
        if _is_blank_row(row):
            continue
        if row[0] in (None, ""):
            continue
        origin_periods.append(str(row[0]))
        row_vals: list[float | None] = []
        for j in range(n_dev):
            col_idx = j + 1
            cell = row[col_idx] if col_idx < len(row) else None
            row_vals.append(_to_float(cell))
        values.append(row_vals)

    if not origin_periods:
        raise ParseError("Pivot tabloda veri satırı yok")

    if not opts.cumulative:
        # Incremental → running sum row by row.
        for row_vals in values:
            running = 0.0
            for j, cell in enumerate(row_vals):
                if cell is None:
                    continue
                running += cell
                row_vals[j] = running

    try:
        return Triangle(
            origin_periods=origin_periods,
            development_periods=development_periods,
            values=values,
            triangle_type=opts.triangle_type,
            origin_granularity=opts.origin_granularity,
            development_granularity=opts.development_granularity,
        )
    except ValueError as e:
        raise ParseError(f"Üçgen doğrulaması başarısız: {e}") from e


_QUARTER_RE = re.compile(r"^\s*(\d{4})\s*[-._/ ]?\s*[qQ]?\s*([1-4])\s*$")
_YEAR_RE = re.compile(r"^\s*(\d{4})\s*$")


def _parse_period(raw: Any, granularity: Granularity) -> tuple[str, int]:
    """Return (label, quarter_rank). Rank is always in quarter units for consistent
    diff across mixed granularities."""
    if isinstance(raw, datetime | date):
        year, month = raw.year, raw.month
        if granularity == Granularity.QUARTERLY:
            q = (month - 1) // 3 + 1
            return f"{year}Q{q}", year * 4 + (q - 1)
        return str(year), year * 4

    if isinstance(raw, int | float):
        if granularity == Granularity.QUARTERLY:
            s = str(int(raw))
            m = _QUARTER_RE.match(s)
            if m:
                year, q = int(m.group(1)), int(m.group(2))
                return f"{year}Q{q}", year * 4 + (q - 1)
            year = int(raw)
            if 1900 <= year <= 2100:
                return f"{year}Q1", year * 4
            raise ValueError(f"çeyreklik sayısal format tanınmadı: {raw}")
        year = int(raw)
        return str(year), year * 4

    s = str(raw).strip()
    if granularity == Granularity.QUARTERLY:
        m = _QUARTER_RE.match(s)
        if m:
            year, q = int(m.group(1)), int(m.group(2))
            return f"{year}Q{q}", year * 4 + (q - 1)
        m = _YEAR_RE.match(s)
        if m:
            year = int(m.group(1))
            return f"{year}Q1", year * 4
        raise ValueError(f"çeyreklik format tanınmadı: {s!r}")

    m = _YEAR_RE.match(s)
    if m:
        year = int(m.group(1))
        return str(year), year * 4
    m = _QUARTER_RE.match(s)
    if m:
        year = int(m.group(1))
        return str(year), year * 4
    raise ValueError(f"yıllık format tanınmadı: {s!r}")


def _normalize(cell: Any) -> str:
    if cell is None:
        return ""
    s = str(cell).strip().lower()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    return s


def _find_match(names: list[str], aliases: tuple[str, ...]) -> int | None:
    for i, n in enumerate(names):
        if n in aliases:
            return i
    for i, n in enumerate(names):
        for a in aliases:
            if a in n:
                return i
    return None


def _is_blank_row(row: list[Any]) -> bool:
    return all(c is None or c == "" for c in row)


def _to_float(cell: Any) -> float | None:
    if cell is None or cell == "":
        return None
    if isinstance(cell, int | float):
        return float(cell)
    try:
        return float(str(cell).replace(",", "."))
    except (TypeError, ValueError):
        return None
