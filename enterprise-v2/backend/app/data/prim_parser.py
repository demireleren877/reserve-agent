"""Prim verisi parser — Branş / Dönem / EP sütunları."""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass

# openpyxl LAZY: startup'ta yüklenmesin (~130ms). Sadece Excel parse edilince yüklenir.
def _has_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


@dataclass
class PrimRecord:
    brans: str
    donem: str   # "2020", "2021" gibi yıl ya da "2020Q1" gibi çeyrek
    ep: float


REQUIRED_FIELDS = ["brans", "donem", "ep"]

_AUTO_ALIASES: dict[str, str] = {
    "brans": "brans", "branch": "brans", "lob": "brans", "line_of_business": "brans",
    "donem": "donem", "yil": "donem", "year": "donem", "period": "donem",
    "accident_year": "donem", "policy_year": "donem",
    "ep": "ep", "earned_premium": "ep", "earnedpremium": "ep",
    "prim": "ep", "kazanilmis_prim": "ep", "kazanilmisprim": "ep",
    "gross_ep": "ep", "net_ep": "ep",
}

_XLSX_MAGIC = b"PK\x03\x04"


def _normalize(name: str) -> str:
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))
    name = name.lower().strip()
    return re.sub(r"[^a-z0-9]+", "_", name).strip("_")


def _auto_suggest(headers: list[str]) -> dict[str, str]:
    suggestion: dict[str, str] = {}
    for h in headers:
        field = _AUTO_ALIASES.get(_normalize(h))
        if field and field not in suggestion:
            suggestion[field] = h
    return suggestion


def _parse_float(s: str) -> float:
    if not s or s.strip() in ("-", ""):
        return 0.0
    s = s.strip().replace("\xa0", "").replace(" ", "")
    if "," in s and "." in s:
        if s.rindex(",") > s.rindex("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _normalize_donem(s: str) -> str:
    """Dönem değerini normalize et: yıl veya yılQ çeyrek."""
    s = str(s).strip()
    # Tam yıl: "2020", "2020.0"
    if re.fullmatch(r"\d{4}\.0", s):
        return s[:4]
    if re.fullmatch(r"\d{4}", s):
        return s
    # yyyyQq formatı
    m = re.fullmatch(r"(\d{4})[Qq]([1-4])", s)
    if m:
        return f"{m.group(1)}Q{m.group(2)}"
    # 6 haneli yyyymm veya yyyyqq
    if re.fullmatch(r"\d{6}", s):
        q = int(s[4:])
        if 1 <= q <= 4:
            return f"{s[:4]}Q{q}"
    return s


def _read_rows_xlsx(content: bytes, sheet_name: str | None) -> tuple[list[str], list[list[str]]]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    data = [[str(c).strip() if c is not None else "" for c in row] for row in rows[1:]]
    return headers, data


def _read_rows_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def inspect_prim_file(content: bytes, filename: str) -> dict:
    """
    Dosyayı incele; sheet listesi, header'lar, preview ve önerilen eşleştirme döner.
    """
    is_xlsx = content[:4] == _XLSX_MAGIC[:4] or filename.lower().endswith((".xlsx", ".xls"))

    if is_xlsx and _has_openpyxl():
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheets: list[str | None] = wb.sheetnames if len(wb.sheetnames) > 1 else [None]
        headers_map: dict[str, list[str]] = {}
        preview_map: dict[str, list[list[str]]] = {}
        suggested_map: dict[str, dict[str, str]] = {}
        for sh in sheets:
            key = sh if sh is not None else "null"
            h, rows = _read_rows_xlsx(content, sh)
            headers_map[key] = h
            preview_map[key] = rows[:5]
            suggested_map[key] = _auto_suggest(h)
        return {
            "sheets": sheets,
            "headers": headers_map,
            "preview": preview_map,
            "suggested_mapping": suggested_map,
        }
    else:
        h, rows = _read_rows_csv(content)
        return {
            "sheets": [None],
            "headers": {"null": h},
            "preview": {"null": rows[:5]},
            "suggested_mapping": {"null": _auto_suggest(h)},
        }


def parse_prim_with_mapping(
    content: bytes,
    filename: str,
    column_mapping: dict[str, str],
    sheet_name: str | None = None,
) -> list[PrimRecord]:
    """
    column_mapping: {field: column_header}  — field ∈ {"brans", "donem", "ep"}
    """
    missing = [f for f in REQUIRED_FIELDS if f not in column_mapping]
    if missing:
        raise ValueError(f"Eksik sütun eşleştirme: {missing}")

    is_xlsx = content[:4] == _XLSX_MAGIC[:4] or filename.lower().endswith((".xlsx", ".xls"))
    if is_xlsx and _has_openpyxl():
        headers, rows = _read_rows_xlsx(content, sheet_name)
    else:
        headers, rows = _read_rows_csv(content)

    col_idx: dict[str, int] = {}
    for field, col_name in column_mapping.items():
        try:
            col_idx[field] = headers.index(col_name)
        except ValueError:
            raise ValueError(f"Sütun bulunamadı: '{col_name}'")

    records: list[PrimRecord] = []
    for line_no, row in enumerate(rows, start=2):
        try:
            brans = row[col_idx["brans"]].strip()
            donem = _normalize_donem(row[col_idx["donem"]])
            ep = _parse_float(row[col_idx["ep"]])
        except (IndexError, ValueError):
            raise ValueError(f"Satır {line_no}: parse hatası")
        if not brans or not donem:
            continue
        records.append(PrimRecord(brans=brans, donem=donem, ep=ep))

    return records
